#! python3
# class_email_notification.py - automatically sends email to parents in the excel file called course_notification.xlsx

# 1. Open the excell file
# 2. Find parent information
# 3. Find course information
# 4. Find student information
# 5. Link student information to the course information
# 6. Send email to parents

import openpyxl, smtplib, sys, datetime
#from time import sleep

wb = openpyxl.load_workbook('course_notification.xlsx')
sheet = wb['student info']

parents = {}
class_info = {}
students = {}
day = datetime.datetime(2020, 6, 1, 8, 0 ,0).strftime('%A, %B, %d, %Y %I:%M:%S %p')

for r in range(2, sheet.max_row + 1):
    parent_name = sheet.cell(row = r, column = 8).value
    parent_email = sheet.cell(row = r, column = 9).value
    parents[parent_name] = parent_email

    course = sheet.cell(row = r, column = 3).value
    account = sheet.cell(row = r, column = 4).value
    link = sheet.cell(row = r, column = 5).value
    clasS_day = sheet.cell(row = r, column = 6).value
    session = sheet.cell(row = r, column = 7).value
    class_info[course] = account, link, clasS_day, session

    gender = sheet.cell(row = r, column = 2).value
    student_name = sheet.cell(row = r, column = 1).value
    
smtpObj = smtplib.SMTP('smtp.gmail.com')
smtpObj.ehlo()
smtpObj.starttls()

for parent_name, parent_email in parents.items():
    for course,(account, link, clasS_day, session) in class_info.items():
        smtpObj.login('tastebolder@gmail.com', sys.argv[1])
        if gender.lower() == 'male':
            gender = 'his'
            body = 'Subject: Tinker Education Online Classes\n\nDear %s,\n\nHope you and your family are doing well and keeping safe while at home. Tinker Education would like to bring to your attention that our online classes will commence  on ', day,' .\n\nOur teachers would like to have an insanely great start of the term with our students. Hence, you are requested to notify your child of ',gender, ' class time. Class details are as follows:\n\n%s\n\nDay: %s\nSession: %s\nCLP:\n%s\nVideo link: %s\n\n%s is requested to keep time of ', gender, ' %s %s\n\nKind regards,\nTinker Desk'%(parent_name, course, clasS_day, session, account, link, student_name, course, session)
        else:
            gender = 'her'
            body = 'Subject: Tinker Education Online Classes\n\nDear %s,\n\nHope you and your family are doing well and keeping safe while at home. Tinker Education would like to bring to your attention that our online classes will commence  on ', day,' .\n\nOur teachers would like to have an insanely great start of the term with our students. Hence, you are requested to notify your child of ',gender, ' class time. Class details are as follows:\n\n%s\n\nDay: %s\nSession: %s\nCLP:\n%s\nVideo link: %s\n\n%s is requested to keep time of ', gender, ' %s %s\n\nKind regards,\nTinker Desk'%(parent_name, course, clasS_day, session, account, link, student_name, course, session)
        print('Sending emails to %s through %s'%(parent_name, parent_email))
        send_email_status = smtpObj.sendmail('tastebolder@gmail.com', parent_email, body)

if send_email_status != {}:
    print('There was an error sending an email to %s through %s'(parent_name, parent_email))
smtpObj.quit()