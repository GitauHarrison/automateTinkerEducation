#! python3
# class_email_notification.py - automatically sends email to parents in the excel file called course_notification.xlsx

# 1. Open the excell file
# 2. Find parent information
# 3. Find course information
# 4. Find student information
# 5. Link student information to the course information
# 6. Send email to parents

import openpyxl, smtplib
from time import sleep

wb = openpyxl.load_workbook('course_notification.xlsx')
sheet = wb['student info']

parent = {}
class_info = {}
students = {}

for r in range(2, sheet.max_row + 1):
    parent_name = sheet.cell(row = r, column = 8).value
    parent_email = sheet.cell(row = r, column = 9).value
    parent[parent_name] = parent_email

    sleep(3)

    course = sheet.cell(row = r, column = 3).value
    account = sheet.cell(row = r, column = 4).value
    link = sheet.cell(row = r, column = 5).value
    clasS_day = sheet.cell(row = r, column = 6).value
    session = sheet.cell(row = r, column = 7).value
    class_info[course] = account, link, clasS_day, session

    sleep(3)

    gender = sheet.cell(row = r, column = 2).value
    if gender == 'male':
        gender = 'his'
    else:
        gender = 'her'
    student_name = sheet.cell(row = r, column = 1).value
